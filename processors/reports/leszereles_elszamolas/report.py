# -*- coding: utf-8 -*-

import os
from collections import defaultdict
import warnings
import settings
from openpyxl import load_workbook
from openpyxl.styles import Font
from processors.reports.base import ReportBase

warnings.simplefilter("ignore")


def chunks(l, n):
    out = []
    for i in range(0, len(l), n):
        out.append(l[i:i + n])
    return out


class LeszerelesElsz(ReportBase):

    def __init__(self, logger):
        super(ReportBase, self).__init__(logger)

    def initialize(self):
        def validate_num(num, min_val, max_val):
            try:
                return int(num) <= max_val and int(num >= min_val)
            except Exception:
                return False

        def validate_y(y):
            return validate_num(y, 2010, 2050)

        def validate_m(m):
            return validate_num(m, 1, 12)

        def validate_d(d):
            return validate_num(d, 1, 31)

        self._sy = self._input(u'(kezdes) Add meg az evet: ', validate_y)
        self._sm = self._input(u'(kezdes) Add meg a honapot: ', validate_m)
        self._sm = self._sm.zfill(2)
        self._sd = self._input(u'(kezdes) Add meg a napot: ', validate_d)
        self._sd = self._sd.zfill(2)

        self._ey = self._input(u'(vege) Add meg az evet: ', validate_y)
        self._em = self._input(u'(vege) Add meg a honapot: ', validate_m)
        self._em = self._em.zfill(2)
        self._ed = self._input(u'(vege) Add meg a napot: ', validate_d)
        self._ed = self._ed.zfill(2)

        self._template_path = os.path.join(os.path.dirname(__file__),
                                           'template.xlsx')
        self._std_font = Font(name='Arial', size=8)

    def _input(self, msg, validator):
        inp = raw_input(msg)
        while not validator(inp):
            print 'Ervenytelen ertek.'
            inp = raw_input(msg)
        return inp

    def _empty_dev(self):
        return {
            'vart_box': [],
            'box': [],
        }

    def _dev_ok(self, dev):
        return dev and dev not in ('0',)

    def _stdize(self, cell):
        """
        Standardize - Applies the standard formatting on a cell.
        """
        cell.font = self._std_font

    def execute(self):
        select_stmt = '''
            SELECT [Felvitel]
            ,[MtAzon]
            ,[JegyAzon]
            ,[uNev]
            ,[tNev]
            ,[tIrsz]
            ,[Utca]
            ,[Vart_Box]
            ,[Vart_Kartya]
            ,[Lezarva]
            ,[Box]
            ,[rlMegjegyzes]
            ,[Kartya]
            ,[mhtNev]
            ,[uMegjegyzes]
            ,[Beveve]
            FROM [leszereles].[dbo].[Ugyfelek]
            LEFT JOIN [leszereles].[dbo].[BoxKartyak] ON [bk_uAzon] = [uAzon]
            LEFT JOIN [leszereles].[dbo].[Telepules] ON [u_tAzon] = [tAzon]
            LEFT JOIN [leszereles].[dbo].[v_MegHiusulasTipusok]
                ON [u_mhtId] = [mhtId]
            LEFT JOIN [leszereles].[dbo].[Raklapok] ON  [rlId] = [bk_rlId]
            WHERE Beveve > '{}-{}-{} 00:00:00.000'
            AND Beveve < '{}-{}-{} 23:59:59.999'
            ORDER BY Felvitel
        '''.format(self._sy, self._sm, self._sd,
                   self._ey, self._em, self._ed)

        self.logger.info('Adatok kikerese folyamatban...')
        conn = self._connect_db(self.LESZ_DB, as_dict=True)
        cursor = conn.cursor()
        cursor.execute(select_stmt)
        report_data = cursor.fetchall()
        conn.close()
        self.logger.info('Kesz')

        workbook = load_workbook(self._template_path)
        worksheet = workbook.get_sheet_by_name('sikeresek')

        # Collect the devices
        self.logger.info('Adatok szamitasa/kiirasa folyamatban...')
        devices = defaultdict(self._empty_dev)

        report_data_cleaned = []
        for row in report_data:
            if 'nincs' in row['uNev'].lower():
                stmt = u'''
                SELECT [Felvitel]
                ,[MtAzon]
                ,[JegyAzon]
                ,[uNev]
                ,[tNev]
                ,[tIrsz]
                ,[Utca]
                ,[Vart_Box]
                ,[Vart_Kartya]
                ,[Lezarva]
                ,[Box]
                ,[rlMegjegyzes]
                ,[Kartya]
                ,[mhtNev]
                ,[uMegjegyzes]
                ,[Beveve]
                FROM [leszereles].[dbo].[Ugyfelek]
                LEFT JOIN [leszereles].[dbo].[BoxKartyak] ON [bk_uAzon] = [uAzon]
                LEFT JOIN [leszereles].[dbo].[Telepules] ON [u_tAzon] = [tAzon]
                LEFT JOIN [leszereles].[dbo].[v_MegHiusulasTipusok]
                    ON [u_mhtId] = [mhtId]
                LEFT JOIN [leszereles].[dbo].[Raklapok] ON  [rlId] = [bk_rlId]
                WHERE Box='{}'
                ORDER BY Felvitel
                '''.format(row['Box'])
                conn = self._connect_db(self.LESZ_DB, as_dict=True)
                cursor = conn.cursor()
                cursor.execute(select_stmt)
                rows = cursor.fetchall()
                conn.close()
                best_row = rows[-1]
                best_row['MtAzon'] = best_row['MtAzon'].strip('xb')
                report_data_cleaned.append(best_row)
            else:
                report_data_cleaned.append(row)

        for row in report_data_cleaned:
            if self._dev_ok(row['Vart_Box']):
                box = {
                    'box': row['Vart_Box'],
                    'kartya': row['Vart_Kartya'] if
                    self._dev_ok(row['Vart_Kartya']) else None
                }
                devices[row['MtAzon']]['vart_box'].append(box)

            if self._dev_ok(row['Box']):
                box = {
                    'box': row['Box'],
                    'kartya': row['Kartya'] if
                    self._dev_ok(row['Kartya']) else None,
                    'tipus': row['rlMegjegyzes']
                }
                devices[row['MtAzon']]['box'].append(box)

        for _, dev_data in devices.iteritems():
            exp_boxes = [d['box'] for d in dev_data['vart_box']]
            for box in dev_data['box']:
                if box['box'] not in exp_boxes:
                    dev_data['vart_box'].append(box)

        # First row is row 3, we skip the header
        row_idx = 3

        already_written = set()
        for row in report_data_cleaned:
            if row['MtAzon'] not in already_written:
                vart_box_rows = chunks(devices[row['MtAzon']]['vart_box'], 3)
                box_rows = chunks(devices[row['MtAzon']]['box'], 3)
                num_rows = max(len(vart_box_rows), len(box_rows))
                for row_num in range(num_rows):
                    try:
                        vart_box_row = vart_box_rows[row_num]
                    except IndexError:
                        vart_box_row = []
                    try:
                        box_row = box_rows[row_num]
                    except IndexError:
                        box_row = []

                    self._write_basic_data(worksheet, row_idx, row)
                    self._write_vart_box_data(worksheet, row_idx, vart_box_row)
                    self._write_box_data(worksheet, row_idx, box_row)

                    row_idx += 1

                already_written.add(row['MtAzon'])
            else:
                continue

        wb_path = os.path.join(settings.REPORT_OUTPUT_DIR,
                               ('leszereles_{}{}{}-{}{}{}.xlsx'
                                ''.format(self._sy, self._sm, self._sd,
                                          self._ey, self._em, self._ed)))

        workbook.save(wb_path)
        self.logger.info('Kiirva ide: {}'.format(wb_path))

    def _write_basic_data(self, ws, row_idx, row):
        # Static field -> column map, the rest is calculated
        field_col_map = {
            'Felvitel': 1,
            'MtAzon': 2,
            'JegyAzon': 3,
            'uNev': 4,
            # 'Beveve': 15,
        }

        ws.row_dimensions[row_idx].height = 12
        for f in field_col_map:
            self._stdize(ws.cell(column=field_col_map[f],
                                 row=row_idx,
                                 value=row[f]))
        address = u'{tNev} {Utca}'.format(**row)
        self._stdize(ws.cell(column=5, row=row_idx, value=address))
        beveve = row['Beveve'].strftime('%Y-%m-%d')
        self._stdize(ws.cell(column=15, row=row_idx, value=beveve))

    def _write_vart_box_data(self, ws, row_idx, data):
        start_col = 7
        for i in range(len(data)):
            col = start_col + i*2
            box = data[i]['box']
            kartya = data[i]['kartya']
            self._stdize(ws.cell(column=col, row=row_idx, value=box))
            self._stdize(ws.cell(column=col+1, row=row_idx, value=kartya))

    def _write_box_data(self, ws, row_idx, data):
        start_col = 16
        for i in range(len(data)):
            col = start_col + i*5
            box = data[i]['box']
            kartya = data[i]['kartya']
            tipus = data[i]['tipus'] or ''
            remote_stbs = ('adb', 'co', 'gcr', 'huawei', 'intek', 'isb', 'x',
                           'sagem')
            remote_stbs_excl = ('sagemcom')
            remote = int(tipus.lower().startswith(remote_stbs) and not
                         tipus.lower().startswith(remote_stbs_excl))
            self._stdize(ws.cell(column=col, row=row_idx, value=box))
            self._stdize(ws.cell(column=col+1, row=row_idx, value=kartya))
            self._stdize(ws.cell(column=col+2, row=row_idx, value=tipus))
            self._stdize(ws.cell(column=col+3, row=row_idx, value=remote))
